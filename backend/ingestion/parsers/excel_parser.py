"""Excel Parser — extracts text from .xlsx/.xls files."""

import io
import openpyxl
from ingestion import TextChunk
from utils.logger import logger


def parse_excel(content: bytes, filename: str) -> list[TextChunk]:
    """Parse an Excel file, converting each sheet's rows to text."""
    chunks = []
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        logger.info(f"Parsing Excel: {filename} ({len(wb.sheetnames)} sheets)")

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue

            # Use first row as headers if it looks like headers
            headers = [str(cell) if cell is not None else "" for cell in rows[0]]
            data_rows = rows[1:] if len(rows) > 1 else rows

            row_texts = []
            for row in data_rows:
                cells = [str(cell) if cell is not None else "" for cell in row]
                if any(c.strip() for c in cells):
                    # Format as "Header: Value" pairs
                    pairs = []
                    for h, v in zip(headers, cells):
                        if v.strip():
                            pairs.append(f"{h}: {v}" if h.strip() else v)
                    row_texts.append(", ".join(pairs))

            if row_texts:
                # Group rows into chunks (50 rows per chunk)
                for i in range(0, len(row_texts), 50):
                    batch = row_texts[i:i + 50]
                    chunks.append(TextChunk(
                        text=f"Sheet: {sheet_name}\n" + "\n".join(batch),
                        page_number=i // 50 + 1,
                        section=f"Sheet: {sheet_name}",
                        metadata={"source": filename, "sheet": sheet_name, "headers": headers},
                    ))

        logger.info(f"Excel parsed: {len(chunks)} chunks")
    except Exception as e:
        logger.error(f"Excel parsing failed for {filename}: {e}")
        raise

    return chunks
